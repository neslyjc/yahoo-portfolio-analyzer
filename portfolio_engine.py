#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MAJ Excel (1 clic) via Yahoo Finance avec yfinance (v8_9)

✅ v8.5 (changement demandé)
- Ajout ligne 19: "Niveau de Risque" (Faible / Moyen / Élevé)
  * Couleurs gérées dans le modèle Excel (mise en forme conditionnelle)
  * Le script remplit la ligne 19 à chaque exécution.
- Conserve v8.4:
  * Tickers source de vérité: "Fonds-Compagnies-suivi" ligne 3 (B→...)
  * Synchronisation automatique de l'onglet "YF_Data" (A1..)
  * Rendements 1/3/5/10 ans = CAGR (annualisés)
  * YTD: si prix du 1er janvier introuvable => "depuis inception"
  * Market Cap (ligne 5) vide pour ETFs: XWD.TO, XTOT.TO, XSP.TO, XQQ.TO
  * Date (B1) = DATE seulement (YYYY-MM-DD)

Modèle attendu:
- Feuille: "Fonds-Compagnies-suivi"
- Ligne 3: tickers en colonnes B→... (jusqu'à cellule vide)
- Lignes fixes (le script écrit aux mêmes numéros de lignes, dont 19)

Prérequis:
  py -m pip install -r requirements_yahoo_v8_8_yfinance.txt
"""

import sys
import math
import os
import datetime as dt
from io import BytesIO
from dataclasses import dataclass
from typing import Optional, List, Tuple

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import yfinance as yf

FILL_UNDERVALUED = PatternFill(fill_type="solid", start_color="A9D18E", end_color="A9D18E")
FILL_NEUTRAL_UNDER = PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE")
FILL_FAIR = PatternFill(fill_type="solid", start_color="E7E6E6", end_color="E7E6E6")
FILL_NEUTRAL_EXP = PatternFill(fill_type="solid", start_color="FCE4D6", end_color="FCE4D6")
FILL_OVERVALUED = PatternFill(fill_type="solid", start_color="F4CCCC", end_color="F4CCCC")

FILL_NEAR_MIN_STRONG = PatternFill(fill_type="solid", start_color="A9D18E", end_color="A9D18E")  # vert
FILL_NEAR_MIN_LIGHT  = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")  # jaune

FILL_NEAR_MAX_STRONG = PatternFill(fill_type="solid", start_color="F4CCCC", end_color="F4CCCC")  # rouge
FILL_NEAR_MAX_LIGHT  = PatternFill(fill_type="solid", start_color="FCE4D6", end_color="FCE4D6")  # orange


ETF_SANS_MARKET_CAP = {"XWD.TO", "XTOT.TO", "XSP.TO", "XQQ.TO"}

# Risque simple (modifiable au besoin)
RISK_MAP = {
    "XWD.TO": "Moyen",   # actions mondiales
    "XTOT.TO": "Moyen",  # actions US total
    "XSP.TO": "Moyen",   # S&P 500
    "XQQ.TO": "Élevé",   # NASDAQ-100 (plus volatil)
    "XCHP.TO": "Élevé",  # secteur semi-conducteurs
    "CGL.TO": "Moyen",   # or
}
DEFAULT_STOCK_RISK = "Élevé"


def now_str() -> str:
    return dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def write_log(log_path: str, msg: str) -> None:
    with open(log_path, "a", encoding="utf-8") as f:
        f.write(f"[{now_str()}] {msg}\n")


def fmt_float(x) -> Optional[float]:
    try:
        if x is None:
            return None
        x = float(x)
        if math.isnan(x) or math.isinf(x):
            return None
        return x
    except Exception:
        return None


def cagr(p0: float, p1: float, years: float) -> Optional[float]:
    if p0 is None or p1 is None or p0 <= 0 or p1 <= 0 or years <= 0:
        return None
    return (p1 / p0) ** (1.0 / years) - 1.0

# =========================
# DEBUT RSI (Relative Strength Index)
# =========================

def compute_rsi(close_prices, period: int = 14):
    if close_prices is None or len(close_prices) < period + 1:
        return None

    gains = []
    losses = []

    for i in range(1, period + 1):
        delta = close_prices[i] - close_prices[i - 1]
        if delta >= 0:
            gains.append(delta)
            losses.append(0)
        else:
            gains.append(0)
            losses.append(abs(delta))

    avg_gain = sum(gains) / period
    avg_loss = sum(losses) / period

    if avg_loss == 0:
        return 100.0

    rs = avg_gain / avg_loss
    rsi = 100 - (100 / (1 + rs))
    return round(rsi, 1)

# =========================
# FIN RSI (Relative Strength Index)
# =========================

@dataclass
class Metrics:
    price: Optional[float] = None
    market_cap_bn: Optional[float] = None
    cagr_10y: Optional[float] = None
    pct_off_52w_high: Optional[float] = None
    ret_1m: Optional[float] = None
    ret_3m: Optional[float] = None
    ret_6m: Optional[float] = None
    ret_ytd_or_inception: Optional[float] = None
    cagr_1y: Optional[float] = None
    cagr_3y: Optional[float] = None
    cagr_5y: Optional[float] = None
    pe: Optional[float] = None
    fwd_pe: Optional[float] = None
    ps: Optional[float] = None
    analyst: Optional[str] = None
    risk: Optional[str] = None
    rsi: Optional[float] = None   # ← AJOUT
    min_12m: Optional[float] = None   # ← AJOUT
    max_12m: Optional[float] = None   # ← AJOUT


def compute_risk(symbol: str) -> str:
    return RISK_MAP.get(symbol, DEFAULT_STOCK_RISK)


def compute_metrics(symbol: str, log_path: str) -> Metrics:
    m = Metrics()
    m.risk = compute_risk(symbol)

    t = yf.Ticker(symbol)

    # Infos (ratios, market cap, analyst)
    info = {}
    try:
        info = t.get_info()
    except Exception as e:
        write_log(log_path, f"[WARN] {symbol}: get_info échoue ({type(e).__name__}: {e})")
        info = {}

    if symbol not in ETF_SANS_MARKET_CAP:
        mc = info.get("marketCap")
        if mc is not None:
            m.market_cap_bn = fmt_float(mc / 1e9)

    m.pe = fmt_float(info.get("trailingPE"))
    m.fwd_pe = fmt_float(info.get("forwardPE"))
    m.ps = fmt_float(info.get("priceToSalesTrailing12Months"))

    rk = info.get("recommendationKey")
    rm = info.get("recommendationMean")
    if isinstance(rk, str) and rk.strip():
        m.analyst = rk.strip().title()
    elif isinstance(rm, (int, float)):
        m.analyst = f"{float(rm):.2f}"

    # History (max pour supporter "depuis inception")
    try:
        hist = t.history(period="max", interval="1d", auto_adjust=True)
    except Exception as e:
        write_log(log_path, f"[WARN] {symbol}: history(max) échoue ({type(e).__name__}: {e})")
        return m
    if hist is None or hist.empty:
        return m

    close = hist["Close"]
    last_date = hist.index[-1].date()
    last_price = float(close.iloc[-1])
    m.price = last_price

    first_date = hist.index[0].date()
    first_price = float(close.iloc[0])

    # =========================
    # DEBUT RSI (Relative Strength Index)
    # =========================
    try:
        closes = close.tail(15).tolist()  # 14 périodes + 1
        m.rsi = compute_rsi(closes)
        # ===== Ligne à supprimer après test ====================
        live_log(f"RSI {ticker}:{rsi}")
        # =========================
    except Exception:
        m.rsi = None
    # =========================
    # FIN RSI (Relative Strength Index)
    # =========================

    def price_on_or_before(target_date: dt.date) -> Optional[Tuple[dt.date, float]]:
        sub = close[close.index.date <= target_date]
        if sub.empty:
            return None
        d = sub.index[-1].date()
        return (d, float(sub.iloc[-1]))

    def simple_return(days: int) -> Optional[float]:
        r = price_on_or_before(last_date - dt.timedelta(days=days))
        if not r:
            return None
        _, p0 = r
        return (last_price / p0) - 1.0

    def annualized_return(years: float) -> Optional[float]:
        days = int(round(years * 365))
        r = price_on_or_before(last_date - dt.timedelta(days=days))
        if not r:
            return None
        d0, p0 = r
        real_years = max((last_date - d0).days / 365.25, 0.0)
        if real_years <= 0:
            return None
        return cagr(p0, last_price, real_years)

    # Rendements court terme (simples)
    m.ret_1m = simple_return(30)
    m.ret_3m = simple_return(91)
    m.ret_6m = simple_return(182)

    # YTD sinon depuis inception
    y0 = dt.date(last_date.year, 1, 1)
    y0_point = price_on_or_before(y0)
    if y0_point:
        _, p_y0 = y0_point
        m.ret_ytd_or_inception = (last_price / p_y0) - 1.0
    else:
        m.ret_ytd_or_inception = (last_price / first_price) - 1.0
        write_log(log_path, f"[INFO] {symbol}: YTD indisponible -> 'depuis inception' ({first_date} -> {last_date})")

    # CAGR 1/3/5/10 ans (annualisés)
    m.cagr_1y = annualized_return(1.0)
    m.cagr_3y = annualized_return(3.0)
    m.cagr_5y = annualized_return(5.0)
    m.cagr_10y = annualized_return(10.0)

    # % sous le plus haut 52 semaines
    cutoff = last_date - dt.timedelta(days=365)
    window = close[close.index.date >= cutoff]
    if not window.empty:
        high52 = float(window.max())
        if high52 > 0:
            m.pct_off_52w_high = (last_price / high52) - 1.0

    # =========================
    # DEBUT Min / Max 12 mois
    # =========================
    cutoff_12m = last_date - dt.timedelta(days=365)
    window_12m = close[close.index.date >= cutoff_12m]

    if not window_12m.empty:
        m.min_12m = round(float(window_12m.min()), 2)
        m.max_12m = round(float(window_12m.max()), 2)
    else:
        m.min_12m = None
        m.max_12m = None
    # =========================
    # FIN Min / Max 12 mois
    # =========================

    return m


ROW_MAP = {
    4: "price",
    5: "market_cap_bn",
    6: "cagr_10y",
    7: "pct_off_52w_high",
    8: "ret_1m",
    9: "ret_3m",
    10: "ret_6m",
    11: "ret_ytd_or_inception",
    12: "cagr_1y",
    13: "cagr_3y",
    14: "cagr_5y",
    15: "pe",
    16: "fwd_pe",
    17: "ps",
    18: "analyst",
    19: "risk",
}


def sync_yf_data(wb, tickers: List[str]) -> None:
    if "YF_Data" in wb.sheetnames:
        yd = wb["YF_Data"]
    else:
        yd = wb.create_sheet("YF_Data")

    yd["A1"] = "Ticker"

    if yd.max_row >= 2:
        yd.delete_rows(2, yd.max_row - 1)

    for i, t in enumerate(tickers, start=2):
        yd.cell(row=i, column=1).value = t

def process_portfolio(uploaded_file, log_callback=None, progress_callback=None):
    log_path = "/tmp/portfolio_log.txt"
    open(log_path, "w", encoding="utf-8").write(f"LOG démarré: {now_str()}\n")

    def live_log(message):
        write_log(log_path, message)
        if log_callback:
            log_callback(message)

    def live_progress(current, total):
        if progress_callback:
            progress_callback(current, total)
    
    wb = load_workbook(uploaded_file)
 
    if "Fonds-Compagnies-suivi" not in wb.sheetnames:
        raise RuntimeError("Feuille 'Fonds-Compagnies-suivi' introuvable.")
    ws = wb["Fonds-Compagnies-suivi"]

    tickers: List[str] = []
    for col in range(2, 500):
        v = ws.cell(row=3, column=col).value
        if v is None or str(v).strip() == "":
            break
        tickers.append(str(v).strip())

    if not tickers:
        raise RuntimeError("Aucun ticker trouvé en ligne 3 (B3→...).")

    live_log(f"Tickers: {tickers}")

    sync_yf_data(wb, tickers)
    live_log("YF_Data synchronisé.")

    ok = 0
    for i, sym in enumerate(tickers):
        col = 2 + i
        try:
            met = compute_metrics(sym, log_path)
            ok += 1
            live_log(f"OK {ticker}")
            
            # ==============================================
            # AJOUT : RULE OF 40 (lignes 20, 21, 22 uniquement)
            # ==============================================

            try:
                info = yf.Ticker(sym).get_info()

                # Croissance annuelle des revenus (%)
                rev_growth = info.get("revenueGrowth")
                rev_growth_pct = round(rev_growth * 100) if isinstance(rev_growth, (int, float)) else None

                # Marge bénéficiaire brute / EBITDA (%)
                ebitda_margin = info.get("ebitdaMargins")
                ebitda_margin_pct = round(ebitda_margin * 100) if isinstance(ebitda_margin, (int, float)) else None

                # Total Rule of 40
                
                if rev_growth_pct is not None and ebitda_margin_pct is not None:
                    rule_of_40 = rev_growth_pct + ebitda_margin_pct
                else:
                    rule_of_40 = None

            except Exception:
                rev_growth_pct = None
                ebitda_margin_pct = None
                rule_of_40 = None

            # Écriture Excel (colonnes B+)
            ws.cell(row=20, column=col).value = "" if rev_growth_pct is None else rev_growth_pct
            ws.cell(row=21, column=col).value = "" if ebitda_margin_pct is None else ebitda_margin_pct
            ws.cell(row=22, column=col).value = "" if rule_of_40 is None else rule_of_40

            # ==================================================
            # FIN AJOUT : RULE OF 40 (lignes 20, 21, 22 uniquement)
            # ==================================================


            # ==================================================
            # DEBUT AJOUT : RSI (lignes 23, 24, 25
            # ==================================================

            ROW_TITLE = 23
            ROW_RANGE = 24
            ROW_EVAL = 25

            plage = ""
            evaluation = ""

            # ===== Ligne 23 : valeur RSI numérique =====
            if met.rsi is None:
                ws.cell(row=ROW_TITLE, column=col).value = ""
            else:
                ws.cell(row=ROW_TITLE, column=col).value = round(met.rsi, 1)
         
            # ===== Lignes 24–25 : interprétation =====

            if met.rsi is None or not isinstance(met.rsi, (int, float)):
                ws.cell(row=ROW_RANGE, column=col).value = ""
                ws.cell(row=ROW_EVAL, column=col).value = ""
            else:
                if met.rsi < 30:
                    plage = "< 30"
                    evaluation = "Undervalued"
                elif 30 <= met.rsi < 50:
                    plage = "30 to 50"
                    evaluation = "Neutral to Undervalued"
                elif met.rsi == 50:
                    plage = "50"
                    evaluation = "Fair Value"
                elif 50 < met.rsi < 70:
                    plage = "50 to 70"
                    evaluation = "Neutral to Expensive"
                else:
                    plage = "70 +"
                    evaluation = "Overvalued"

                ws.cell(row=ROW_RANGE, column=col).value = plage
                ws.cell(row=ROW_EVAL, column=col).value = evaluation

                cell_eval = ws.cell(row=ROW_EVAL, column=col)

                if evaluation == "Undervalued":
                    cell_eval.fill = FILL_UNDERVALUED

                elif evaluation == "Neutral to Undervalued":
                    cell_eval.fill = FILL_NEUTRAL_UNDER

                elif evaluation == "Fair Value":
                    cell_eval.fill = FILL_FAIR

                elif evaluation == "Neutral to Expensive":
                    cell_eval.fill = FILL_NEUTRAL_EXP

                elif evaluation == "Overvalued":
                    cell_eval.fill = FILL_OVERVALUED

                else:
                    cell_eval.fill = PatternFill()  # aucune couleur


            # ==================================================
            # FIN AJOUT : RSI (lignes 23, 24, 25 uniquement)
            # ==================================================


            # ==================================================
            # DEBUT AJOUT : Min / Max 12 mois (lignes 26, 27)
            # ==================================================

            ROW_MIN_12M = 26
            ROW_MAX_12M = 27

            ws.cell(row=ROW_MIN_12M, column=col).value = (
                "" if met.min_12m is None else met.min_12m
            )

            ws.cell(row=ROW_MAX_12M, column=col).value = (
                "" if met.max_12m is None else met.max_12m
            )

            # ==================================================
            # FIN AJOUT : Min / Max 12 mois
            # ==================================================


            # ==================================================
            # DEBUT AJOUT VISUEL : Proximité Min / Max 12 mois
            # ==================================================

            price = met.price
            min_12m = met.min_12m
            max_12m = met.max_12m

            cell_min = ws.cell(row=ROW_MIN_12M, column=col)
            cell_max = ws.cell(row=ROW_MAX_12M, column=col)

            # Reset couleur (important si rerun)
            cell_min.fill = PatternFill()
            cell_max.fill = PatternFill()

            if isinstance(price, (int, float)):

                # ----- Proche du MIN -----
                if isinstance(min_12m, (int, float)):
                    if price <= min_12m * 1.05:
                        cell_min.fill = FILL_NEAR_MIN_STRONG
                elif price <= min_12m * 1.10:
                        cell_min.fill = FILL_NEAR_MIN_LIGHT

            # ----- Proche du MAX -----
            if isinstance(max_12m, (int, float)):
                if price >= max_12m * 0.95:
                    cell_max.fill = FILL_NEAR_MAX_STRONG
                elif price >= max_12m * 0.90:
                      cell_max.fill = FILL_NEAR_MAX_LIGHT

            # ==================================================
            # FIN AJOUT VISUEL : Proximité Min / Max 12 mois
            # ==================================================



            # ==================================================
            # DEBUT AJOUT : SCORE INDICATION D'ACHAT (ligne 28)
            # ==================================================

            ROW_SCORE = 28

            def clamp(x):
                return max(0, min(100, x))

            score_components = []

            # ---- 1) % sous 52W high ----
            if isinstance(met.pct_off_52w_high, (int, float)):
                pct_down = abs(met.pct_off_52w_high) * 100
                score_52w = clamp(100 - (pct_down * 2.5))
                score_components.append(score_52w)

            # ---- 2) Position dans plage 12M ----
            if all(isinstance(v, (int, float)) for v in [price, min_12m, max_12m]):
                if max_12m > min_12m:
                    score_range = ((price - min_12m) / (max_12m - min_12m)) * 100
                    score_components.append(clamp(score_range))

            # ---- 3) Trailing PE ----
            if isinstance(met.pe, (int, float)):
                pe = met.pe
                score_pe = clamp((pe / 60) * 100)
                score_components.append(score_pe)

            # ---- 4) Rule of 40 ----
            if isinstance(rule_of_40, (int, float)):
                if rule_of_40 >= 60:
                    score_r40 = 0
                elif rule_of_40 <= 0:
                    score_r40 = 100
                else:
                    score_r40 = clamp(100 - (rule_of_40 * 1.5))
                score_components.append(score_r40)

            # ---- 5) RSI ----
            if isinstance(met.rsi, (int, float)):
                score_rsi = clamp(met.rsi)
                score_components.append(score_rsi)

            # ---- 6) Compression PE ----
            if isinstance(met.pe, (int, float)) and isinstance(met.fwd_pe, (int, float)):
                if met.fwd_pe < met.pe:
                    score_forward = 20
                else:
                    score_forward = 80
                score_components.append(score_forward)

           
            # ---- Score final ----
            if len(score_components) >= 2:
                final_score = round(sum(score_components) / len(score_components), 1)
            else:
                final_score = None

            cell_score = ws.cell(row=ROW_SCORE, column=col)
            cell_score.value = "" if final_score is None else final_score


            # ----- Couleur -----
            if isinstance(final_score, (int, float)):

                if final_score <= 25:
                    cell_score.fill = PatternFill(fill_type="solid", start_color="70AD47", end_color="70AD47")
                elif final_score <= 50:
                    cell_score.fill = PatternFill(fill_type="solid", start_color="A9D18E", end_color="A9D18E")
                elif final_score <= 70:
                    cell_score.fill = PatternFill(fill_type="solid", start_color="FFD966", end_color="FFD966")
                elif final_score <= 85:
                    cell_score.fill = PatternFill(fill_type="solid", start_color="F4B183", end_color="F4B183")
                else:
                    cell_score.fill = PatternFill(fill_type="solid", start_color="F4CCCC", end_color="F4CCCC")

            # ==================================================
            # FIN AJOUT : SCORE INDICATION D'ACHAT
            # ==================================================

           
        except Exception as e:
            write_log(log_path, f"[WARN] {sym}: {type(e).__name__}: {e} (valeurs laissées vides)")
            continue

        for row, attr in ROW_MAP.items():
            ws.cell(row=row, column=col).value = getattr(met, attr)

    ws["B1"] = dt.date.today().strftime("%Y-%m-%d")

    # out = xlsx_path.replace(".xlsx", "_MAJ.xlsx")
    # wb.save(out)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    write_log(log_path, f"Terminé. {ok}/{len(tickers)} tickers traités.")

    return output
    
